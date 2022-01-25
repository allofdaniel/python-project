# python selenium example

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time  # 시간차 위해 적용
import urllib.request
import os
import openpyxl

options = webdriver.ChromeOptions()
# options.add_argument('headless')  # headless
# options.add_argument('window-size=1920x1080')  # headless
# options.add_argument("disable-gpu")  # headless
options.add_experimental_option("excludeSwitches", ["enable-logging"])
driver = webdriver.Chrome(options=options)
wb = openpyxl.Workbook()

# 포인트 읽기
f = open("포인트.txt", 'r', encoding='utf8')
lines = f.readlines(int())

# 대중교통
for idx_org_, org_ in enumerate(lines):
    for idx_des_, des_ in enumerate(lines):
        if idx_org_ != idx_des_:
            driver.get(
                "https://map.naver.com/v5/directions/{0},%EC%B6%9C%EB%B0%9C%EC%A7%80,,/{1},%EB%8F%84%EC%B0%A9%EC%A7%80,,ADDRESS_POI/-/transit?c=14110045.8089916,4510631.0916025,10,0,0,0,dh".format(org_, des_))
            try:
                time.sleep(1)
                time_p = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-pubtransit.item_summary.selected.ng-star-inserted > div:nth-child(1) > strong > readable-duration")
                walktime_p = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-pubtransit.item_summary.selected.ng-star-inserted > div:nth-child(2) > span:nth-child(2) > readable-duration")
                cost_p = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-pubtransit.item_summary.selected.ng-star-inserted > div:nth-child(2) > span:nth-child(3)")
            except:
                text = open('대중교통.txt', 'a')
                text.write("zone{} to zone{} 결과없음\n".format(
                    idx_org_, idx_des_))
                pass
            else:
                a_p = time_p.text
                b_p = walktime_p.text
                c_p = cost_p.text
                text = open('대중교통.txt', 'a')
                text.write("zone{} to zone{}".format(idx_org_, idx_des_))
                text.write(' 총 {}, 도보 {}, 비용 {}\n'.format(a_p, b_p, c_p))
            # finally:
            #     text.close()
            #     driver.quit()
        else:
            text = open('대중교통.txt', 'a')
            text.write("zone{} to zone{} 동일경로\n".format(idx_org_, idx_des_))
            text.close()


# 자동차  class="link_tab car"
for idx_org_, org_ in enumerate(lines):
    for idx_des_, des_ in enumerate(lines):
        if idx_org_ != idx_des_:
            driver.get(
                "https://map.naver.com/v5/directions/{0},%EC%84%9C%EC%9A%B8%ED%8A%B9%EB%B3%84%EC%8B%9C%20%EC%A2%85%EB%A1%9C%EA%B5%AC%20%EC%B2%AD%EC%9A%B4%ED%9A%A8%EC%9E%90%EB%8F%99,,/{1},%EC%84%9C%EC%9A%B8%ED%8A%B9%EB%B3%84%EC%8B%9C%20%EC%A2%85%EB%A1%9C%EA%B5%AC%20%EA%B0%80%ED%9A%8C%EB%8F%99,,/-/car?c=14135277.4956642,4520807.1271698,15,0,0,0,dh".format(org_, des_))
            try:
                time.sleep(1)
                fnltime_car = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-car.item_summary.selected.ng-star-inserted > div.summary_box > strong > readable-duration")
                distance_car = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-car.item_summary.selected.ng-star-inserted > div.summary_box > span.summary_text > readable-distance")
                tollfee = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-car.item_summary.selected.ng-star-inserted > div.route_text_box.fee > span:nth-child(1)")
                gasprice = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-car.item_summary.selected.ng-star-inserted > div.route_text_box.fee > span:nth-child(2) > span")
                taxicost = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-car.item_summary.selected.ng-star-inserted > div.route_text_box.fee > span:nth-child(3) > span")
            except:
                text = open('자동차.txt', 'a')
                text.write("zone{} to zone{} 결과없음\n".format(
                    idx_org_, idx_des_))
                pass
            else:
                a_c = fnltime_car.text
                b_c = distance_car.text
                c_c = tollfee.text
                d_c = gasprice.text
                e_c = taxicost.text
                text = open('자동차.txt', 'a')
                text.write("zone{} to zone{}".format(idx_org_, idx_des_))
                text.write(' 총 {}, 거리 {}, {}, 주유비 {}, 택시비 {}\n'.format(
                    a_c, b_c, c_c, d_c, e_c))
            # finally:
            #     text.close()
            #     driver.quit()
        else:
            text = open('자동차.txt', 'a')
            text.write("zone{} to zone{} 동일경로\n".format(idx_org_, idx_des_))
            text.close()

# 도보 link_tab walk
for idx_org_, org_ in enumerate(lines):
    for idx_des_, des_ in enumerate(lines):
        if idx_org_ != idx_des_:
            driver.get(
                "https://map.naver.com/v5/directions/{0},%EC%84%B1%EB%82%A8%EC%8B%9C%EC%B2%AD,11627831,PLACE_POI/{1},%EC%96%91%EC%82%B0%EC%8B%9C%EC%B2%AD,11628231,PLACE_POI/-/walk?c=14022253.5341211,4351060.4882194,6,0,0,0,dh".format(org_, des_))
            try:
                time.sleep(1)
                fnltime_wlk = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-walking.item_summary.selected.ng-star-inserted > div.summary_box > strong > readable-duration")
                distance_wlk = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-walking.item_summary.selected.ng-star-inserted > div.summary_box > span.summary_text > readable-distance")
                path_wlk = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-walking.item_summary.selected.ng-star-inserted > div.route_box.ng-star-inserted > div > span > directions-facilities")
            except:
                text = open('도보.txt', 'a')
                text.write("zone{} to zone{} 결과없음\n".format(
                    idx_org_, idx_des_))
                pass
            else:
                a_w = fnltime_wlk.text
                b_w = distance_wlk.text
                c_w = path_wlk.text
                text = open('도보.txt', 'a')
                text.write("zone{} to zone{}".format(idx_org_, idx_des_))
                text.write(' 총 {}, 거리 {}, 경로 {}\n'.format(a_w, b_w, c_w))
            # finally:
            #     text.close()
            #     driver.quit()
        else:
            text = open('도보.txt', 'a')
            text.write("zone{} to zone{} 동일경로\n".format(idx_org_, idx_des_))
            text.close()

# 자전거 link_tab bike
for idx_org_, org_ in enumerate(lines):
    for idx_des_, des_ in enumerate(lines):
        if idx_org_ != idx_des_:
            driver.get(
                "https://map.naver.com/v5/directions/{0},%EC%84%9C%EC%9A%B8%ED%8A%B9%EB%B3%84%EC%8B%9C%20%EC%A2%85%EB%A1%9C%EA%B5%AC%20%EC%B2%AD%EC%9A%B4%ED%9A%A8%EC%9E%90%EB%8F%99,,/{1},%EC%84%9C%EC%9A%B8%ED%8A%B9%EB%B3%84%EC%8B%9C%20%EC%A2%85%EB%A1%9C%EA%B5%AC%20%EA%B0%80%ED%9A%8C%EB%8F%99,,/-/bike?c=14134599.1170381,4520558.7068279,15,0,0,0,dh".format(org_, des_))
            try:
                time.sleep(1)
                fnltime_bike = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-bike > div.summary_box > strong > readable-duration")
                distance_bike = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-bike > div.summary_box > span.summary_text > readable-distance")
                route_bike = driver.find_element_by_css_selector(
                    "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-bike > div.route_box.ng-star-inserted > div")
                # path_bike = driver.find_element_by_css_selector(
                #     "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-bike > div.route_box.ng-star-inserted > div.route_text_box.route_bike.ng-star-inserted > span > directions-facilities")
            except:
                text = open('자전거.txt', 'a')
                text.write("zone{} to zone{} 결과없음\n".format(
                    idx_org_, idx_des_))
                pass
            else:
                a_b = fnltime_bike.text
                b_b = distance_bike.text
                c_b = route_bike.text
                # d_b = path_bike.text
                text = open('자전거.txt', 'a')
                text.write("zone{} to zone{}".format(idx_org_, idx_des_))
                text.write(' 총 {}, 거리 {}, 경로 {}\n'.format(
                    a_b, b_b, c_b))
            # finally:
            #     text.close()
            #     driver.quit()
        else:
            text = open('자전거.txt', 'a')
            text.write("zone{} to zone{} 동일경로\n".format(idx_org_, idx_des_))
            text.close()
# for i in range(len((lines)*len(lines))):
#     driver.quit()

# 오류검정과정 넣어야함. 같으면 pass.
# excel로 만들어야함.


# 워크북(엑셀파일)을 새로 만듭니다.


# sheet1 = wb.active # sheet1은 활성화되 있는 시트를 선택
# sheet1.title = "1st sheet" # sheet1의 시트이름 변경
# sheet2 = wb.create_sheet("2nd sheet") # 새로운 시트 만들고 sheet2에 저장

# for i in range(1, 10):
#     sheet1.cell(row=i, column=1).value = i
#     sheet2.cell(row=1, column=i).value = i

# 워크북(엑셀파일)을 원하는 이름으로 저장합니다.
