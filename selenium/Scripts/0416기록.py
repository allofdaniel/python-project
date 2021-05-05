# # python selenium example

# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# # until 쓰기위한 #https://velog.io/@log327/Python-Selenium-Explicit-Waits-%EC%82%AC%EC%9A%A9%ED%95%98%EA%B8%B0
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.keys import Keys
# import requests  # 자원 얻기위해
# import time  # 시간차 위해 적용
# import urllib.request
# import os
# options = webdriver.ChromeOptions()
# options.add_experimental_option("excludeSwitches", ["enable-logging"])
# driver = webdriver.Chrome(options=options)
# # 포인트 읽기
# f = open("포인트.txt", 'r', encoding='utf8')
# lines = f.readlines(int())
# for idx_org_, org_ in enumerate(lines):
#     for idx_des_, des_ in enumerate(lines):
#         if idx_org_ != idx_des_:
#             driver.get(
#                 "https://map.naver.com/v5/directions/{0},%EC%B6%9C%EB%B0%9C%EC%A7%80,,/{1},%EB%8F%84%EC%B0%A9%EC%A7%80,,ADDRESS_POI/-/transit?c=14110045.8089916,4510631.0916025,10,0,0,0,dh".format(org_, des_))
#             try:
#                 fnltime = WebDriverWait(driver, 5).until(EC.visibility_of(
#                     (By.css_selector, "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-pubtransit.item_summary.selected.ng-star-inserted > div:nth-child(1) > strong > readable-duration")))
#                 walktime = WebDriverWait(driver, 5).until(EC.visibility_of(
#                     (By.css_selector, "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-pubtransit.item_summary.selected.ng-star-inserted > div:nth-child(2) > span:nth-child(2) > readable-duration")))
#                 fnlcost = WebDriverWait(driver, 5).until(EC.visibility_of(
#                     (By.css_selector, "container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-pubtransit.item_summary.selected.ng-star-inserted > div:nth-child(2) > span:nth-child(3)")))
#             except:
#                 # text = open('거리속도.txt', 'a')
#                 # text.write("zone{} to zone{} 결과없음 ".format(idx_org_, idx_des_))
#                 pass
#             else:
#                 a = fnltime.text
#                 b = walktime.text
#                 c = fnlcost.text
#                 text = open('거리속도.txt', 'a')
#                 text.write("zone{} to zone{}".format(idx_org_, idx_des_))
#                 text.write(' 총{}, 도보{}, 비용{}\n'.format(a, b, c))
#             # finally:
#             #     text.close()
#             #     driver.quit()
#         else:
#             text = open('거리속도.txt', 'a')
#             text.write("zone{} to zone{} 동일경로\n".format(idx_org_, idx_des_))
#             text.close()


# # 오류검정과정 넣어야함. 같으면 pass.
# # excel로 만들어야함.
# #               driver.get(
# #                     "https://map.naver.com/v5/directions/{0},%EC%B6%9C%EB%B0%9C%EC%A7%80,,/{1},%EB%8F%84%EC%B0%A9%EC%A7%80,,ADDRESS_POI/-/transit?c=14110045.8089916,4510631.0916025,10,0,0,0,dh".format(org_, des_))
# #                 time.sleep(3)
# #                 fnltime = driver.find_element_by_css_selector(
# #                     "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-pubtransit.item_summary.selected.ng-star-inserted > div:nth-child(1) > strong > readable-duration")
# #                 walktime = driver.find_element_by_css_selector(
# #                     "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-pubtransit.item_summary.selected.ng-star-inserted > div:nth-child(2) > span:nth-child(2) > readable-duration")
# #                 fnlcost = driver.find_element_by_css_selector(
# #                     "#container > shrinkable-layout > div > directions-layout > directions-result > div.main > directions-summary-list > directions-hover-scroll > div > div > directions-summary-item-pubtransit.item_summary.selected.ng-star-inserted > div:nth-child(2) > span:nth-child(3)")
# #                 a = fnltime.text
# #                 b = walktime.text
# #                 c = fnlcost.text

# # 위의 코드 구현 x
# beautiful soup 으로도 잘 안될것이다.

# open xl
# # openpyxl을 가져옵니다.
import openpyxl

# 워크북(엑셀파일)을 새로 만듭니다.
wb = openpyxl.Workbook()

# sheet1은 활성화되 있는 시트를 선택
sheet1 = wb.active
# sheet1의 시트이름 변경
sheet1.title = "1st sheet"

# 새로운 시트 만들고 sheet2에 저장
sheet2 = wb.create_sheet("2nd sheet")

# sheet1과 sheet2에 동시에 데이터 쓰기
for i in range(1, 10):
    sheet1.cell(row=i, column=1).value = i
    sheet2.cell(row=1, column=i).value = i

# 워크북(엑셀파일)을 원하는 이름으로 저장합니다.
wb.save('test.xlsx')
